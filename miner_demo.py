import os
import time

import pandas as pd
import requests

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

from openpyxl import load_workbook
from openpyxl.styles import Alignment,Border, Side


url = "https://filfox.info/en"
html = requests.get(url).text
print(html)



wd_main = webdriver.Chrome()
wd = webdriver.Chrome()
wd_main.get(url)
time.sleep(2)

xpath1 = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div[7]/div[2]/table/tbody/tr[1]/td[2]/span/span/span/a"
xpath2 = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div[7]/div[2]/table/tbody/tr[2]/td[2]/span/span/span/a"
xpath3 = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div[7]/div[2]/table/tbody/tr[3]/td[2]/span/span/span/a"

xpath_prefix = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div[7]/div[2]/table/tbody/tr["
xpath_suffix = "]/td[2]/span/span/span/a"

miner_list = []
for i in range(1,11):
    miner = wd_main.find_element(By.XPATH, xpath_prefix+str(i)+xpath_suffix).text

    print(f"miner{i} = {miner}")

    miner_url_prefix = "https://filfox.info/en/address/"

    miner_url = miner_url_prefix + miner

    wd.get(miner_url)

    #如果网页有错误，重新加载
    while "error" in wd.title:
        wd.get(miner_url)

    pow_xpath = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div/div[3]/div[2]/div[2]/div/div[1]/p[1]"

    time.sleep(2)
    pow = wd.find_element(By.XPATH, pow_xpath).text.strip().split(" ")[0]
    print(f"pow = {pow}")

    win_count_xpath = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div/div[3]/div[2]/div[2]/div/div[2]/div"
    win_count = wd.find_element(By.XPATH, win_count_xpath).text.split(":")[1].strip()
    print(f"win_count = {win_count}")

    total_reward_xpath = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div/div[3]/div[2]/div[2]/div/div[3]/p[1]"
    total_reward = wd.find_element(By.XPATH, total_reward_xpath).text.split(":")[1].strip().split(" ")[0]
    print(f"total_reward = {total_reward}")

    locked_rewards_xpath = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div/div[3]/div[2]/div[1]/div/div[2]/p[5]"
    locked_rewards = wd.find_element(By.XPATH, locked_rewards_xpath).text.split(":")[1].strip().split(" ")[0]
    print(f"locked_rewards = {locked_rewards}")

    day_blocks_xpath = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div/div[4]/div[2]/div[2]/div[1]"
    day_blocks = wd.find_element(By.XPATH, day_blocks_xpath).text.split(":")[1].strip()
    print(f"day_blocks = {day_blocks}")

    day_reword_xpath = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div/div[4]/div[2]/div[2]/p"
    day_reward = wd.find_element(By.XPATH, day_reword_xpath).text.split(":")[1].strip().split(" ")[0]
    print(f"day_reward = {day_reward}")

    day_lucky_xpath = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div/div[4]/div[2]/div[3]/p[2]"
    day_lucky = wd.find_element(By.XPATH, day_lucky_xpath).text.split(":")[1].strip()
    print(f"day_lucky = {day_lucky}")

    label2_xpath = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div/div[4]/div[1]/div/div/label[2]"

    wait = WebDriverWait(wd, 5)
    week_label = wait.until(EC.element_to_be_clickable((By.XPATH, label2_xpath)))
    week_label.click()

    time.sleep(2)
    week_blocks_xpath = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div/div[4]/div[2]/div[2]/div[1]"
    week_blocks = wd.find_element(By.XPATH, week_blocks_xpath).text.split(":")[1].strip()
    print(f"week_blocks = {week_blocks}")

    week_reword_xpath = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div/div[4]/div[2]/div[2]/p"
    week_reward = wd.find_element(By.XPATH, week_reword_xpath).text.split(":")[1].strip().split(" ")[0]
    print(f"week_reward = {week_reward}")

    week_lucky_xpath = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div/div[4]/div[2]/div[3]/p[2]"
    week_lucky = wd.find_element(By.XPATH, week_lucky_xpath).text.split(":")[1].strip()
    print(f"week_lucky = {week_lucky}")

    label3_xpath = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div/div[4]/div[1]/div/div/label[3]"

    wait = WebDriverWait(wd, 5)
    month_label = wait.until(EC.element_to_be_clickable((By.XPATH, label3_xpath)))
    month_label.click()

    time.sleep(2)
    month_blocks_xpath = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div/div[4]/div[2]/div[2]/div[1]"
    month_blocks = wd.find_element(By.XPATH, month_blocks_xpath).text.split(":")[1].strip()
    print(f"month_blocks = {month_blocks}")

    month_reword_xpath = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div/div[4]/div[2]/div[2]/p"
    month_reward = wd.find_element(By.XPATH, month_reword_xpath).text.split(":")[1].strip().split(" ")[0]
    print(f"month_reward = {month_reward}")

    month_lucky_xpath = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div/div[4]/div[2]/div[3]/p[2]"
    month_lucky = wd.find_element(By.XPATH, month_lucky_xpath).text.split(":")[1].strip()
    print(f"month_lucky = {month_lucky}")

    label4_xpath = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div/div[4]/div[1]/div/div/label[4]"

    wait = WebDriverWait(wd, 5)
    year_label = wait.until(EC.element_to_be_clickable((By.XPATH, label4_xpath)))
    year_label.click()

    time.sleep(2)
    year_blocks_xpath = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div/div[4]/div[2]/div[2]/div[1]"
    year_blocks = wd.find_element(By.XPATH, year_blocks_xpath).text.split(":")[1].strip()
    print(f"year_blocks = {year_blocks}")

    year_reword_xpath = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div/div[4]/div[2]/div[2]/p"
    year_reward = wd.find_element(By.XPATH, year_reword_xpath).text.split(":")[1].strip().split(" ")[0]
    print(f"year_reward = {year_reward}")

    year_lucky_xpath = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div/div[4]/div[2]/div[3]/p[2]"
    year_lucky = wd.find_element(By.XPATH, year_lucky_xpath).text.split(":")[1].strip()
    print(f"year_lucky = {year_lucky}")
    print("\n")

    miner_data = {
        "节点号":miner,
        "算力(P)":pow,
        "总出块":win_count,
        "总奖励(FIL)":total_reward,
        "锁仓":locked_rewards,
        "出块(天)":day_blocks,
        "奖励(天)":day_reward,
        "幸运值(天)":day_lucky,
        "出块(周)": week_blocks,
        "奖励(周)": week_reward,
        "幸运值(周)": week_lucky,
        "出块(月)": month_blocks,
        "奖励(月)": month_reward,
        "幸运值(月)": month_lucky,
        "出块(年)": year_blocks,
        "奖励(年)": year_reward,
        "幸运值(年)": year_lucky,
    }
    miner_list.append(miner_data)

wd_main.close()
wd.close()

df = pd.DataFrame(miner_list)

time_now = time.strftime("%Y-%m-%d(%H-%M-%S)",time.localtime())
excel_filename = time_now+"-miners_data.xlsx"

df.to_excel(excel_filename,index=False,engine='openpyxl')

print(f"数据已写入Excel文件:{os.getcwd()}\\{excel_filename}")

wb = load_workbook(excel_filename)

ws = wb.active

# 设置列宽
ws.column_dimensions['A'].width = 11
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 13
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 12
ws.column_dimensions['I'].width = 10
ws.column_dimensions['J'].width = 13
ws.column_dimensions['K'].width = 12
ws.column_dimensions['L'].width = 10
ws.column_dimensions['M'].width = 14
ws.column_dimensions['N'].width = 12
ws.column_dimensions['O'].width = 10
ws.column_dimensions['P'].width = 14
ws.column_dimensions['Q'].width = 12

# 设置所有单元格的对齐方式为居中
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# 设置有数据区域内的单元格边框
# 定义边框样式
border = Border(left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))

# 设置A1:D10区域的边框
for row in ws['A1':'Q11']:
    for cell in row:
        cell.border = border

wb.save(excel_filename)



