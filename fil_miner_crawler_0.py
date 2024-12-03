import re
import time

from lxml import etree
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Alignment,Border, Side

# 谷歌浏览器驱动程序的路径
# chrome_driver_path = "C:\\Program Files\\Google\\Chrome\\chromedriver.exe"
chrome_driver_path = "D:\\google\\chrome\\chromedriver.exe"
# 浏览器选项
options = Options()
# options.add_argument("--headless")  # Temporarily removed for debugging

# 谷歌浏览器服务
service = Service(chrome_driver_path)

# 谷歌浏览器驱动句柄 handle
driver = webdriver.Chrome(service=service, options=options)

# 循环抓取每个节点的数据：
# 总算力值:power,
# 总出块数量：blocks_total
# FIL总产出数量：fils_total
# FIL锁仓数量：fils_locked
# 幸运值：24小时，7天，30天，1年
# 出块数量：24小时，7天，30天，1年
# FIL产量：24小时，7天，30天，1年

# 矿工节点号列表：一共4个节点
miners = ["f01992032", "f01992563", "f01996719", "f01996817"]
# miners = ["f01083914", "f01697248", "f02208475", "f01889512", "f02216186", "f01245980","f01084913","f02131881","f02062851","f02274508"]


data_list = []
for miner in miners:
    try:
        url = "https://filfox.info/en/address/" + miner
        driver.get(url)
        html_content = driver.page_source
        # 解析html标签
        html = etree.HTML(html_content)

        print("节点号：" + miner)
        # 取总算力值
        print("总算力值：")
        divs = html.xpath("/html/body/div[1]/div/div/div[1]/div[1]/div/div[3]/div[2]/div[2]/div/div[1]/p[1]/text()")
        print(divs)
        power = divs[0][:6].strip()
        print(power)

        # 取总出块数量
        print("总出块数量：")
        divs = html.xpath("/html/body/div[1]/div/div/div[1]/div[1]/div/div[3]/div[2]/div[2]/div/div[2]/div/text()[2]")
        blocks_total = divs[0][2:].strip()
        print(blocks_total)

        # 取FIL总产出数量
        print("FIL总产出数量：")
        divs = html.xpath("/html/body/div[1]/div/div/div[1]/div[1]/div/div[3]/div[2]/div[2]/div/div[3]/p[1]/text()")
        fils_total = divs[0][15:24].strip()
        print(fils_total)

        # 取FIL锁仓数量
        print("FIL锁仓数量：")
        divs = html.xpath("/html/body/div[1]/div/div/div[1]/div[1]/div/div[3]/div[2]/div[1]/div/div[2]/p[5]/text()")
        fils_locked = divs[0][17:27].strip()
        print(fils_locked)

        wait = WebDriverWait(driver, 10)  # Increased timeout for debugging
        selector = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div/div[4]/div[1]/div/div/label[1]"
        hours_24_tab = wait.until(EC.element_to_be_clickable((By.XPATH, selector)))
        hours_24_tab.click()

        # Wait for the updated element
        updated_selector = "div.mx-8.border.border-background.rounded-sm.p-4"
        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, updated_selector)))

        html_content = driver.page_source
        # 解析
        html = etree.HTML(html_content)

        # 取当天幸运值
        # divs = html.xpath("/html/body/div[1]/div/div/div[1]/div[1]/div/div[4]/div[2]/div[3]/p[2]/text()")
        print("幸运值（24小时）")
        divs = html.xpath("/html/body/div/div/div/div[1]/div[1]/div/div[4]/div[2]/div[3]/p[2]/text()[2]")
        lucky_value = divs[0][2:].strip()
        print(lucky_value)
        lucky_value_24h = lucky_value

        # 取当天出块数量
        print("出块数量（24小时）")
        divs = html.xpath("/html/body/div[1]/div/div/div[1]/div[1]/div/div[4]/div[2]/div[2]/div[1]/text()[2]")
        blocks = divs[0][2:].strip()
        print(blocks)
        blocks_24h = blocks

        # 取当天产出数量FIL
        print("FIL产量（24小时）")
        divs = html.xpath("/html/body/div[1]/div/div/div[1]/div[1]/div/div[4]/div[2]/div[2]/p/text()")
        fils = divs[0]
        fils = fils.split(':')[1].split(' ')[1]
        fils_24h = fils

        print("\n")

        # 获取并保存当前幸运值的文本
        initial_blocks_xpath = "/html/body/div[1]/div/div/div[1]/div[1]/div/div[4]/div[2]/div[2]/div[1]"
        initial_blocks_text = driver.find_element(By.XPATH, initial_blocks_xpath).text

        # 点击第2个标签
        selector = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div/div[4]/div[1]/div/div/label[2]"
        hours_7day_tab = wait.until(EC.element_to_be_clickable((By.XPATH, selector)))
        hours_7day_tab.click()

        # 等待幸运值的文本发生变化
        wait.until(
            lambda driver: driver.find_element(By.XPATH, initial_blocks_xpath).text != initial_blocks_text
        )

        html_content = driver.page_source
        # 解析
        html = etree.HTML(html_content)

        # 取7天幸运值
        print("幸运值（7天）")
        divs = html.xpath("/html/body/div/div/div/div[1]/div[1]/div/div[4]/div[2]/div[3]/p[2]/text()[2]")
        lucky_value = divs[0][2:].strip()
        print(lucky_value)
        lucky_value_7d = lucky_value


        # 取7天出块数量
        print("出块数量（7天）")
        divs = html.xpath("/html/body/div[1]/div/div/div[1]/div[1]/div/div[4]/div[2]/div[2]/div[1]/text()[2]")
        blocks = divs[0][2:].strip()
        print(blocks)
        blocks_7d = blocks

        # 取7天产出数量FIL
        print("FIL产量（7天）")
        divs = html.xpath("/html/body/div[1]/div/div/div[1]/div[1]/div/div[4]/div[2]/div[2]/p/text()")
        fils = divs[0]
        fils = fils.split(':')[1].split(' ')
        print(fils[1])
        fils_7d = fils[1]

        print("\n")

        # 点击第3个标签
        selector = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div/div[4]/div[1]/div/div/label[3]"
        hours_30day_tab = wait.until(EC.element_to_be_clickable((By.XPATH, selector)))
        hours_30day_tab.click()

        day30_block_text = driver.find_element(By.XPATH, initial_blocks_xpath).text

        # 等待幸运值的文本发生变化
        wait.until(
            lambda driver: driver.find_element(By.XPATH, initial_blocks_xpath).text != day30_block_text
        )

        html_content = driver.page_source
        # 解析
        html = etree.HTML(html_content)

        # 取30天幸运值
        print("幸运值（30天）")
        divs = html.xpath("/html/body/div/div/div/div[1]/div[1]/div/div[4]/div[2]/div[3]/p[2]/text()[2]")
        lucky_value = divs[0][2:].strip()
        print(lucky_value)
        lucky_value_30d = lucky_value

        # 取30天出块数量
        print("出块数量（30天）")
        divs = html.xpath("/html/body/div[1]/div/div/div[1]/div[1]/div/div[4]/div[2]/div[2]/div[1]/text()[2]")
        blocks = divs[0][2:].strip()
        print(blocks)
        blocks_30d = blocks

        # 取30天产出数量FIL
        print("FIL产量（30天）")
        divs = html.xpath("/html/body/div[1]/div/div/div[1]/div[1]/div/div[4]/div[2]/div[2]/p/text()")
        fils = divs[0]
        fils = fils.split(':')[1].split(' ')
        print(fils[1])
        fils_30d = fils[1]
        print("\n")

        # 点击第4个标签
        selector = "//*[@id=\"__layout\"]/div/div[1]/div[1]/div/div[4]/div[1]/div/div/label[4]"
        hours_30day_tab = wait.until(EC.element_to_be_clickable((By.XPATH, selector)))
        hours_30day_tab.click()

        day30_block_text = driver.find_element(By.XPATH, initial_blocks_xpath).text

        # 等待幸运值的文本发生变化
        wait.until(
            lambda driver: driver.find_element(By.XPATH, initial_blocks_xpath).text != day30_block_text
        )

        html_content = driver.page_source
        # 解析
        html = etree.HTML(html_content)

        # 取幸运值（1年）
        print("幸运值（1年）")
        divs = html.xpath("/html/body/div/div/div/div[1]/div[1]/div/div[4]/div[2]/div[3]/p[2]/text()[2]")
        lucky_value = divs[0][2:].strip()
        print(lucky_value)
        lucky_value_1y = lucky_value

        # 取出块数量（1年）
        print("出块数量（1年）")
        divs = html.xpath("/html/body/div[1]/div/div/div[1]/div[1]/div/div[4]/div[2]/div[2]/div[1]/text()[2]")
        blocks = divs[0][2:].strip()
        print(blocks)
        blocks_1y = blocks

        # 取FIL产量（1年）
        print("FIL产量（1年）")
        divs = html.xpath("/html/body/div[1]/div/div/div[1]/div[1]/div/div[4]/div[2]/div[2]/p/text()")
        fils = divs[0]
        fils = fils.split(':')[1].split(' ')
        print(fils[1])
        fils_1y = fils[1]
        print("\n")

        data = {
            '节点号': miner,
            '总算力值': power,
            '总出块': blocks_total,
            '总产量(FIL)': fils_total,
            '锁仓量(FIL)': fils_locked,
            '幸运值(1天)': lucky_value_24h,
            '幸运值(7天)': lucky_value_7d,
            '幸运值(30天)': lucky_value_30d,
            '幸运值(1年)': lucky_value_1y,
            '出块(1天)': blocks_24h,
            '出块(7天)': blocks_7d,
            '出块(30天)': blocks_30d,
            '出块(1年)': blocks_1y,
            'FIL产量(1天)': fils_24h,
            'FIL产量(7天)': fils_7d,
            'FIL产量(30天)': fils_30d,
            'FIL产量(1年)': fils_1y,
        }
        data_list.append(data)


    except TimeoutException as e:
        print(f"等待元素超时. 未能找到元素: {selector}")
        print(f"异常信息: {e.msg}")
        print("当前页面源码:", driver.page_source[:2000])  # 打印部分页面源码

driver.quit()

# 将数据列表转换成DataFrame
df = pd.DataFrame(data_list)

# 定义Excel文件名
time_now = time.strftime("%Y-%m-%d(%H-%M-%S)", time.localtime())
excel_filename = "E:\\atoshi\\fildata\\"+time_now+"-miners_data.xlsx"

# 将DataFrame写入Excel文件
df.to_excel(excel_filename, index=False, engine='openpyxl')

print(f'数据已写入Excel文件：{excel_filename}')

wb = load_workbook(excel_filename)

ws = wb.active

# 设置列宽
ws.column_dimensions['A'].width = 11
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 13
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 12
ws.column_dimensions['I'].width = 10
ws.column_dimensions['J'].width = 13
ws.column_dimensions['K'].width = 12
ws.column_dimensions['L'].width = 10
ws.column_dimensions['M'].width = 14
ws.column_dimensions['N'].width = 12
ws.column_dimensions['O'].width = 10
ws.column_dimensions['P'].width = 14
ws.column_dimensions['Q'].width = 12

# 设置所有单元格的对齐方式为居中
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# 设置有数据区域内的单元格边框
# 定义边框样式
border = Border(left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))

# 设置A1:D10区域的边框
for row in ws['A1':'Q11']:
    for cell in row:
        cell.border = border

wb.save(excel_filename)